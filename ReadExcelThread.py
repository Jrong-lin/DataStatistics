import openpyxl
from PyQt5.QtCore import QThread, pyqtSignal


class cReadExcelThread(QThread):
    signalAppendItem = pyqtSignal(list)
    def __init__(self,strFile):
        super(cReadExcelThread, self).__init__()
        self.m_strFile = strFile
        workbook = openpyxl.load_workbook(self.m_strFile)
        self.booksheet = workbook.active
        self.m_maxRow = self.booksheet.max_row
    def run(self):
        self.listCurData = []
        for col in range(1, self.booksheet.max_column + 1):
            if "店铺名称" == self.booksheet.cell(1,col).value:
                self.shopNameIndex = col
                continue
            elif "商品编码" in self.booksheet.cell(1,col).value and "线上" in self.booksheet.cell(1,col).value:
                self.commodityCodeIndex = col
                continue
            elif "规格" in self.booksheet.cell(1,col).value and "线上" in self.booksheet.cell(1,col).value:
                self.specificationIndex = col
                continue
            elif "线上订单状态" == self.booksheet.cell(1,col).value:
                self.OrderStatusOlIndex = col
                continue
            elif "明细状态" in self.booksheet.cell(1,col).value:
                self.detailStatusIndex = col
                continue
            elif "实付总计" in self.booksheet.cell(1,col).value:
                self.totalRealPayIndex = col
                print(self.totalRealPayIndex)
                continue
            elif "数量" in self.booksheet.cell(1,col).value:
                self.numberIndex = col
                continue

        for row in range(2, self.booksheet.max_row + 1):
            self.listCurData.clear()
            self.listCurData.append(self.booksheet.cell(row,self.shopNameIndex).value)
            self.listCurData.append(self.booksheet.cell(row, self.commodityCodeIndex).value)
            self.listCurData.append(self.booksheet.cell(row, self.specificationIndex).value)
            self.listCurData.append(self.booksheet.cell(row, self.OrderStatusOlIndex).value)
            self.listCurData.append(self.booksheet.cell(row, self.detailStatusIndex).value)
            self.listCurData.append(self.booksheet.cell(row, self.totalRealPayIndex).value)
            self.listCurData.append(self.booksheet.cell(row, self.numberIndex).value)
            self.signalAppendItem.emit(self.listCurData)

    def CurFileMaxRow(self):
        return self.m_maxRow
